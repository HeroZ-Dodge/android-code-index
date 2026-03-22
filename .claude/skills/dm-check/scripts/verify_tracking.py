#!/usr/bin/env python3
"""
埋点验证工具 v2
- 支持 spec JSON / Excel 输入
- 输出 HTML 验证报告（概览表格 + JSON 详情 + 关注字段标红）
- 触发行为分析
"""

import argparse
import json
import os
import re
import subprocess
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from html import escape as esc

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ==================== Constants ====================

ELK_URL = "http://api.elk.x.netease.com:9550/a19_drpf_clientlog-*/_search"
ELK_TOKEN = "20b6e03e3f8a48c2b2019d7e47d286bb"
TZ_EAST8 = timezone(timedelta(hours=8))

COMMON_FIELDS = {
    "f", "t", "msg_key", "log_sid", "platform", "f_ver", "info_append_way",
    "user_uid", "user_name", "role_name", "gameid", "gameimei", "deviceid",
    "app_ver", "system", "channel", "urs", "network",
    "active_community_id", "active_community_name", "active_page_struc",
    "session_id", "refer_action_id", "refer_action_param",
    "content",
    "proxima_meta", "@timestamp", "@version", "_score",
    "deviceid_v2", "game_group_id", "identity_authentic_info",
    "user_sex", "mac_addr", "user_account_type", "cguid",
    "role_list", "role_server", "role_grade", "user_birth", "user_location",
    "timestamp", "sid", "log_source", "log_source_inf", "log_token",
    "is_mkey", "latest_channel",
}

# ==================== Time Parsing ====================

def parse_time_range(time_range_str):
    now = datetime.now(TZ_EAST8)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    tr = time_range_str.strip().lower() if time_range_str else "today"

    presets = {
        "today": (today_start, now),
        "今天": (today_start, now),
        "yesterday": (today_start - timedelta(days=1), today_start - timedelta(seconds=1)),
        "昨天": (today_start - timedelta(days=1), today_start - timedelta(seconds=1)),
        "last1h": (now - timedelta(hours=1), now),
        "最近1小时": (now - timedelta(hours=1), now),
        "last24h": (now - timedelta(hours=24), now),
        "最近24小时": (now - timedelta(hours=24), now),
        "last3d": (now - timedelta(days=3), now),
        "最近3天": (now - timedelta(days=3), now),
        "last7d": (now - timedelta(days=7), now),
        "最近7天": (now - timedelta(days=7), now),
    }

    if tr in presets:
        start, end = presets[tr]
    else:
        try:
            day = datetime.strptime(tr, "%Y-%m-%d").replace(tzinfo=TZ_EAST8)
            start, end = day, day.replace(hour=23, minute=59, second=59)
        except ValueError:
            print(f"WARNING: Cannot parse time range '{time_range_str}', using today", file=sys.stderr)
            start, end = today_start, now

    return int(start.timestamp() * 1000), int(end.timestamp() * 1000)


# ==================== ELK ====================

def query_elk(uid, event_names, start_millis, end_millis):
    should_clauses = [{"match_phrase": {"f": evt}} for evt in event_names]
    query_body = {
        "query": {
            "bool": {
                "must": [
                    {"match_phrase": {"user_uid": {"query": uid}}},
                    {"bool": {"should": should_clauses, "minimum_should_match": 1}},
                    {"range": {"@timestamp": {"gte": start_millis, "lte": end_millis, "format": "epoch_millis"}}},
                ]
            }
        },
        "size": 10000,
        "sort": [{"@timestamp": {"order": "desc"}}],
    }

    result = subprocess.run(
        ["curl", "-s", "-H", f"ELK-AUTH-TOKEN:{ELK_TOKEN}",
         "-H", "Content-Type:application/json", "-XPOST", ELK_URL,
         "-d", json.dumps(query_body, ensure_ascii=False)],
        capture_output=True, text=True, timeout=30,
    )

    if result.returncode != 0:
        print(f"ERROR: curl failed: {result.stderr}", file=sys.stderr)
        return []

    try:
        resp = json.loads(result.stdout)
    except json.JSONDecodeError:
        print(f"ERROR: Invalid JSON response: {result.stdout[:200]}", file=sys.stderr)
        return []

    if "error" in resp:
        print(f"ERROR: ELK error: {json.dumps(resp['error'], ensure_ascii=False)[:300]}", file=sys.stderr)
        return []

    hits = resp.get("hits", {})
    total = hits.get("total", {}).get("value", 0)
    print(f"  ELK returned {total} hits (fetched {len(hits.get('hits', []))})", file=sys.stderr)
    return hits.get("hits", [])


def parse_elk_records(hits):
    records = []
    for hit in hits:
        source = hit.get("_source", {})
        event_name = source.get("f")
        if not event_name:
            continue
        biz_fields = {}
        content_str = source.get("content")
        if isinstance(content_str, str):
            try:
                inner = json.loads(content_str)
                if isinstance(inner, dict):
                    for k, v in inner.items():
                        if k not in COMMON_FIELDS:
                            biz_fields[k] = v
            except (json.JSONDecodeError, TypeError):
                pass
        for k, v in source.items():
            if k not in COMMON_FIELDS:
                biz_fields[k] = v
        ts = source.get("@timestamp", "")
        records.append({"event_name": event_name, "biz_fields": biz_fields, "raw": source, "timestamp": ts})
    return records


# ==================== Logcat ====================

LOGCAT_PATTERN = re.compile(r'\d+/send:\s*(\{.+\})\s*$')
LOGCAT_TIME_PATTERN = re.compile(r'^(\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}\.\d{3})\s+')


def read_logcat(logcat_path):
    records = []
    with open(logcat_path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            m = LOGCAT_PATTERN.search(line)
            if not m:
                continue
            try:
                obj = json.loads(m.group(1))
            except json.JSONDecodeError:
                continue

            event_name = obj.get("f")
            if not event_name:
                continue

            biz_fields = {}
            content_str = obj.get("content")
            if isinstance(content_str, str):
                try:
                    inner = json.loads(content_str)
                    if isinstance(inner, dict):
                        for k, v in inner.items():
                            if k not in COMMON_FIELDS:
                                biz_fields[k] = v
                except (json.JSONDecodeError, TypeError):
                    pass
            for k, v in obj.items():
                if k not in COMMON_FIELDS:
                    biz_fields[k] = v

            tm = LOGCAT_TIME_PATTERN.match(line)
            ts = tm.group(1) if tm else ""

            records.append({"event_name": event_name, "biz_fields": biz_fields, "raw": obj, "timestamp": ts})
    return records


def capture_logcat():
    import tempfile
    tmp_path = os.path.join(tempfile.gettempdir(), "dm-check-logcat-dump.txt")
    try:
        with open(tmp_path, "w") as outf:
            proc = subprocess.run(
                ["adb", "logcat", "-d", "-v", "time",
                 "GLLogManager:D", "*:S"],
                stdout=outf, stderr=subprocess.PIPE, text=True, timeout=60,
            )
        if proc.returncode != 0:
            print(f"ERROR: adb logcat failed: {proc.stderr}", file=sys.stderr)
            return []
        return read_logcat(tmp_path)
    except subprocess.TimeoutExpired:
        print("WARNING: adb logcat timed out, parsing partial output", file=sys.stderr)
        if os.path.exists(tmp_path):
            return read_logcat(tmp_path)
        return []


# ==================== Excel Parsing ====================

def detect_header(row_cells):
    values = [str(c.value).strip() if c.value else "" for c in row_cells]
    return "事件名" in values


def build_column_map(row_cells):
    col_map = {}
    for cell in row_cells:
        if not cell.value:
            continue
        val = str(cell.value).strip()
        idx = cell.column - 1
        mapping = {
            "事件名": "event_name", "事件中文名": "event_cn_name",
            "字段名": "field_name", "字段id": "field_id", "类型": "field_type",
            "描述": "description", "触发规则": "trigger_rule",
        }
        if val in mapping:
            col_map[mapping[val]] = idx
        elif val in ("备注", "字段内容"):
            col_map["remark"] = idx
    return col_map


def is_section_title(row_cells):
    non_empty = [c for c in row_cells if c.value is not None and str(c.value).strip()]
    if len(non_empty) <= 2:
        first_val = str(non_empty[0].value).strip() if non_empty else ""
        if first_val and not re.match(r'^[a-z_]+$', first_val):
            return True
    return False


def get_cell_value(row_cells, col_map, key):
    idx = col_map.get(key)
    if idx is None or idx >= len(row_cells):
        return None
    val = row_cells[idx].value
    return str(val).strip() if val is not None else None


def extract_expected_values(remark):
    if not remark:
        return []
    values = []

    # "预计记录：\n259：xxx\n249：yyy" → ["259", "249"]
    m = re.search(r'预计记录[：:]\s*(.*)', remark, re.DOTALL)
    if m:
        body = m.group(1)
        nums = re.findall(r'(\d+)\s*[：:]', body)
        if nums:
            return nums

    # "枚举值（看过、收藏、点赞）" or "枚举值：全部、已看完、未看完"
    m = re.search(r'枚举值[（(：:]\s*([^）)]+)', remark)
    if m:
        parts = re.split(r'[、,，/]', m.group(1).strip())
        return [p.strip() for p in parts if p.strip()]

    # "类型（DEFAULT-默认，COMIC-漫剧）" → ["DEFAULT", "COMIC"]
    pairs = re.findall(r'([A-Z_]+)\s*[-—]\s*[\u4e00-\u9fff]', remark)
    if len(pairs) >= 2:
        return pairs

    # "取值: a/b/c" or "取值：a、b、c"
    m = re.search(r'取值[：:]\s*(.+)', remark)
    if m:
        parts = re.split(r'[/、,，]', m.group(1).strip())
        return [p.strip() for p in parts if p.strip()]

    # "1: 是, 0: 否" or "1-xxx 0-yyy"
    pairs = re.findall(r'(\d+)\s*[：:]\s*\S+', remark)
    if len(pairs) >= 2:
        return pairs

    # "固定值：xxx"
    m = re.search(r'固定值[：:\s]+(\S+)', remark)
    if m:
        return [m.group(1)]

    return values


def parse_sheet(ws, sheet_name):
    events = {}
    col_map = None
    current_event = None
    for row in ws.iter_rows():
        cells = list(row)
        values = [c.value for c in cells]
        if all(v is None for v in values):
            continue
        if detect_header(cells):
            col_map = build_column_map(cells)
            current_event = None
            continue
        if col_map is None:
            continue
        if is_section_title(cells):
            current_event = None
            continue
        evt_name = get_cell_value(cells, col_map, "event_name")
        if evt_name and re.match(r'^[a-z_][a-z0-9_]*$', evt_name):
            evt_cn = get_cell_value(cells, col_map, "event_cn_name") or ""
            desc = get_cell_value(cells, col_map, "description") or ""
            trigger = get_cell_value(cells, col_map, "trigger_rule") or ""
            current_event = evt_name
            if current_event not in events:
                events[current_event] = {
                    "name": current_event,
                    "cn_name": evt_cn,
                    "description": desc,
                    "trigger_rule": trigger,
                    "sheet": sheet_name,
                    "fields": {},
                }
        field_id = get_cell_value(cells, col_map, "field_id")
        if field_id and current_event and current_event in events:
            field_name_raw = get_cell_value(cells, col_map, "field_name") or ""
            is_new = "本期新增" in field_name_raw
            field_name = re.sub(r'[（(]本期新增[）)]', '', field_name_raw).strip()
            field_type = get_cell_value(cells, col_map, "field_type") or ""
            remark = get_cell_value(cells, col_map, "remark") or ""
            extracted = extract_expected_values(remark)
            events[current_event]["fields"][field_id] = {
                "field_id": field_id,
                "field_name": field_name,
                "field_type": field_type,
                "remark": remark,
                "extracted_values": extracted,
                "is_new": is_new,
            }
    return events


def parse_excel(excel_path, target_sheets=None):
    if openpyxl is None:
        print("ERROR: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
        sys.exit(1)
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    all_events = {}
    sheets = target_sheets if target_sheets else wb.sheetnames
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            print(f"WARNING: Sheet '{sheet_name}' not found, skipping", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        events = parse_sheet(ws, sheet_name)
        for evt_name, evt_info in events.items():
            if evt_name in all_events:
                all_events[evt_name]["fields"].update(evt_info["fields"])
            else:
                all_events[evt_name] = evt_info
    return all_events


def excel_to_parse_json(excel_events):
    result = {"events": {}}
    for evt_name, evt_info in excel_events.items():
        fields = {}
        for fid, fdef in evt_info["fields"].items():
            fields[fid] = {
                "name": fdef["field_name"],
                "type": fdef["field_type"],
                "remark": fdef["remark"],
                "extracted_values": fdef.get("extracted_values", []),
                "is_new": fdef.get("is_new", False),
            }
        change_type = "modify" if "修改" in evt_info.get("sheet", "") else "new"
        result["events"][evt_name] = {
            "cn_name": evt_info["cn_name"],
            "trigger_rule": evt_info.get("trigger_rule", ""),
            "description": evt_info.get("description", ""),
            "change_type": change_type,
            "sheet": evt_info.get("sheet", ""),
            "fields": fields,
        }
    return result


# ==================== Spec Loading ====================

def load_spec(spec_path):
    with open(spec_path, "r", encoding="utf-8") as f:
        return json.load(f)


# ==================== Trigger Analysis ====================

TRIGGER_MAP = [
    (r'^clk_', "click", "用户点击操作"),
    (r'^exp_', "exposure", "元素曝光（元素可见时触发）"),
    (r'^exposure_', "exposure", "元素曝光（元素可见时触发）"),
    (r'^page_view$', "page_view", "页面曝光（页面从不可见变为可见时触发）"),
    (r'^page_', "page", "页面相关行为"),
    (r'^slide_', "slide", "用户滑动操作"),
    (r'^search_', "search", "搜索行为"),
    (r'^play_', "play", "播放行为"),
]


def infer_trigger(event_name, trigger_rule=""):
    for pattern, ttype, desc in TRIGGER_MAP:
        if re.match(pattern, event_name):
            if trigger_rule:
                return ttype, f"{desc}（{trigger_rule}）"
            return ttype, desc
    if trigger_rule:
        return "custom", trigger_rule
    return "unknown", "未知触发方式"


# ==================== Verification ====================

PLACEHOLDER_VALUES = {"-9999", -9999, "-9999.0"}


def is_placeholder(val):
    if val is None:
        return False
    if isinstance(val, (dict, list)):
        return False
    if isinstance(val, (int, float)) and val == -9999:
        return True
    if isinstance(val, str) and val.strip() in ("-9999", ""):
        return True
    return False


def verify(spec, records):
    events = spec.get("events", {})
    grouped = defaultdict(list)
    for r in records:
        grouped[r["event_name"]].append(r)

    results = []
    for evt_name, evt_info in events.items():
        actual = grouped.get(evt_name, [])
        triggered = len(actual) > 0
        doc_fields = evt_info.get("fields", {})
        trigger_rule = evt_info.get("trigger_rule", "")
        trigger_type, trigger_desc = infer_trigger(evt_name, trigger_rule)

        # Collect actual values per field
        field_values = defaultdict(list)
        field_seen = defaultdict(set)
        all_field_keys = set()

        for rec in actual:
            for k, v in rec["biz_fields"].items():
                all_field_keys.add(k)
                ser = json.dumps(v, ensure_ascii=False, sort_keys=True) if v is not None else "null"
                if ser not in field_seen[k]:
                    field_seen[k].add(ser)
                    field_values[k].append(v)

        # Build field results
        field_results = []
        for fid, fdef in doc_fields.items():
            present = fid in all_field_keys
            actual_vals = field_values.get(fid, [])
            expected_vals = fdef.get("expected_values", [])
            ftype = fdef.get("type", "")

            status = "pass"
            if not present:
                status = "missing"
            else:
                # Type check for json fields
                if ftype.lower() == "json" and actual_vals:
                    sample = actual_vals[0]
                    if isinstance(sample, str):
                        try:
                            json.loads(sample)
                        except (json.JSONDecodeError, TypeError):
                            status = "type_mismatch"

                # Placeholder check
                if status == "pass" and actual_vals:
                    ph_count = sum(1 for v in actual_vals if is_placeholder(v))
                    if ph_count == len(actual_vals):
                        status = "abnormal"
                    elif ph_count > 0:
                        status = "partial_abnormal"

            # Expected value coverage
            coverage = {}
            if expected_vals and actual_vals:
                actual_strs = [str(v) for v in actual_vals]
                for ev in expected_vals:
                    coverage[ev] = any(ev == av for av in actual_strs)
                if status == "pass" and not all(coverage.values()):
                    status = "partial_coverage"

            # Sub-field verification for JSON-type fields
            sub_field_results = []
            sub_fields_def = fdef.get("sub_fields", {})
            if ftype.lower() == "json" and sub_fields_def and present:
                sf_values = defaultdict(list)
                sf_seen = defaultdict(set)
                sf_all_keys = set()

                for av in actual_vals:
                    obj = av
                    if isinstance(obj, str):
                        try:
                            obj = json.loads(obj)
                        except (json.JSONDecodeError, TypeError):
                            continue
                    if isinstance(obj, dict):
                        for k, v in obj.items():
                            sf_all_keys.add(k)
                            ser = json.dumps(v, ensure_ascii=False, sort_keys=True) if v is not None else "null"
                            if ser not in sf_seen[k]:
                                sf_seen[k].add(ser)
                                sf_values[k].append(v)

                for sfid, sfdef in sub_fields_def.items():
                    sf_present = sfid in sf_all_keys
                    sf_actual = sf_values.get(sfid, [])
                    sf_expected = sfdef.get("expected_values", [])

                    sf_status = "pass"
                    if not sf_present:
                        sf_status = "missing"
                    elif sf_actual:
                        ph_count = sum(1 for v in sf_actual if is_placeholder(v))
                        if ph_count == len(sf_actual):
                            sf_status = "abnormal"
                        elif ph_count > 0:
                            sf_status = "partial_abnormal"

                    sf_coverage = {}
                    if sf_expected and sf_actual:
                        sf_actual_strs = [str(v) for v in sf_actual]
                        for ev in sf_expected:
                            sf_coverage[ev] = any(ev == av for av in sf_actual_strs)
                        if sf_status == "pass" and not all(sf_coverage.values()):
                            sf_status = "partial_coverage"

                    sub_field_results.append({
                        "field_id": sfid,
                        "parent_field_id": fid,
                        "name": sfdef.get("name", ""),
                        "type": sfdef.get("type", "string"),
                        "expected_values": sf_expected,
                        "actual_values": sf_actual,
                        "present": sf_present,
                        "status": sf_status,
                        "coverage": sf_coverage,
                        "description": sfdef.get("description", ""),
                        "is_new": sfdef.get("is_new", False),
                    })

            field_results.append({
                "field_id": fid,
                "name": fdef.get("name", ""),
                "type": ftype,
                "expected_values": expected_vals,
                "actual_values": actual_vals,
                "present": present,
                "status": status,
                "coverage": coverage,
                "description": fdef.get("description", ""),
                "is_new": fdef.get("is_new", False),
                "sub_field_results": sub_field_results,
            })

        # Unique combinations of documented fields
        combos = []
        seen_combos = set()
        doc_fids = list(doc_fields.keys())
        for rec in actual:
            combo = {}
            for fid in doc_fids:
                combo[fid] = rec["biz_fields"].get(fid)
            # Also include extra biz fields
            for k, v in rec["biz_fields"].items():
                if k not in combo:
                    combo[k] = v
            key = json.dumps(combo, ensure_ascii=False, sort_keys=True)
            if key not in seen_combos:
                seen_combos.add(key)
                combos.append(combo)

        # Raw records (limit to 50)
        raw_list = [rec["biz_fields"] for rec in actual[:50]]

        # Overall status
        has_field_issue = any(f["status"] != "pass" for f in field_results)
        has_sub_field_issue = any(
            sf["status"] != "pass"
            for f in field_results
            for sf in f.get("sub_field_results", [])
        )
        if not triggered:
            overall = "not_triggered"
        elif has_field_issue or has_sub_field_issue:
            overall = "has_issues"
        else:
            overall = "pass"

        # Collect highlight keys (include sub-field keys for JSON highlighting)
        all_doc_keys = set(doc_fields.keys())
        for fid, fdef in doc_fields.items():
            for sfid in fdef.get("sub_fields", {}).keys():
                all_doc_keys.add(sfid)

        results.append({
            "event_name": evt_name,
            "cn_name": evt_info.get("cn_name", ""),
            "change_type": evt_info.get("change_type", ""),
            "triggered": triggered,
            "count": len(actual),
            "trigger_type": trigger_type,
            "trigger_desc": trigger_desc,
            "overall": overall,
            "fields": field_results,
            "combinations": combos,
            "raw_records": raw_list,
            "doc_field_ids": all_doc_keys,
        })

    return results


# ==================== HTML Report ====================

def render_json_html(obj, highlight_keys=None, indent=0, _parent_highlighted=False):
    hl = highlight_keys or set()
    pad = "  " * indent
    pad1 = "  " * (indent + 1)

    if isinstance(obj, dict):
        if not obj:
            return '{}'
        parts = ['{']
        items = list(obj.items())
        for i, (k, v) in enumerate(items):
            comma = ',' if i < len(items) - 1 else ''
            is_hl = k in hl
            key_cls = 'jhl' if is_hl else 'jk'
            val_html = render_json_html(v, hl, indent + 1, is_hl)
            parts.append(f'{pad1}<span class="{key_cls}">"{esc(str(k))}"</span>: {val_html}{comma}')
        parts.append(f'{pad}}}')
        return '\n'.join(parts)
    elif isinstance(obj, list):
        if not obj:
            return '[]'
        parts = ['[']
        for i, item in enumerate(obj):
            comma = ',' if i < len(obj) - 1 else ''
            val_html = render_json_html(item, hl, indent + 1, _parent_highlighted)
            parts.append(f'{pad1}{val_html}{comma}')
        parts.append(f'{pad}]')
        return '\n'.join(parts)
    elif isinstance(obj, str):
        cls = 'jhv' if _parent_highlighted else 'js'
        return f'<span class="{cls}">"{esc(obj)}"</span>'
    elif isinstance(obj, bool):
        cls = 'jhv' if _parent_highlighted else 'jb'
        return f'<span class="{cls}">{str(obj).lower()}</span>'
    elif isinstance(obj, (int, float)):
        cls = 'jhv' if _parent_highlighted else 'jn'
        return f'<span class="{cls}">{obj}</span>'
    elif obj is None:
        cls = 'jhv' if _parent_highlighted else 'jnl'
        return f'<span class="{cls}">null</span>'
    else:
        return f'<span>{esc(str(obj))}</span>'


def status_badge(status):
    labels = {
        "pass": ("通过", "bp"),
        "missing": ("缺失", "bf"),
        "type_mismatch": ("类型错误", "bf"),
        "abnormal": ("异常值", "bf"),
        "partial_abnormal": ("部分异常", "bw"),
        "partial_coverage": ("值未全覆盖", "bw"),
        "not_triggered": ("未触发", "bm"),
        "has_issues": ("有异常", "bf"),
    }
    label, cls = labels.get(status, (status, "bm"))
    return f'<span class="{cls}">{label}</span>'


def generate_html(results, spec, source_desc, total_records, output_path):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    title = spec.get("title", "埋点验证报告")

    total = len(results)
    passed = sum(1 for r in results if r["overall"] == "pass")
    failed = sum(1 for r in results if r["overall"] == "has_issues")
    not_triggered = sum(1 for r in results if r["overall"] == "not_triggered")

    html = [f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>{esc(title)}</title>
<style>
:root{{--p:#10b981;--f:#ef4444;--w:#f59e0b;--bg:#f8fafc;--c:#fff;--bd:#e2e8f0;--t:#1e293b;--t2:#64748b}}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',system-ui,sans-serif;background:var(--bg);color:var(--t);line-height:1.6;padding:24px}}
.ct{{max-width:1200px;margin:0 auto}}
h1{{font-size:22px;margin-bottom:8px}}
.meta{{color:var(--t2);font-size:13px;margin-bottom:24px}}
.meta span{{margin-right:16px}}
.sm{{display:flex;gap:12px;margin-bottom:24px;flex-wrap:wrap}}
.sc{{background:var(--c);border:1px solid var(--bd);border-radius:8px;padding:14px 20px;flex:1;min-width:130px}}
.sc .lb{{font-size:11px;color:var(--t2);text-transform:uppercase;letter-spacing:.5px}}
.sc .vl{{font-size:26px;font-weight:700}}
.sc.sp .vl{{color:var(--p)}}.sc.sf .vl{{color:var(--f)}}.sc.sw .vl{{color:var(--w)}}.sc.sn .vl{{color:var(--t2)}}
h2{{font-size:17px;margin:24px 0 12px;padding-bottom:8px;border-bottom:2px solid var(--bd)}}
.ot{{width:100%;border-collapse:collapse;margin-bottom:24px;background:var(--c);border-radius:8px;overflow:hidden;border:1px solid var(--bd);font-size:13px}}
.ot th{{background:#f1f5f9;padding:10px 14px;text-align:left;font-weight:600;color:var(--t2);border-bottom:2px solid var(--bd)}}
.ot td{{padding:8px 14px;border-bottom:1px solid var(--bd)}}
.ot tr:last-child td{{border-bottom:none}}
.ot tr:hover{{background:#f8fafc}}
.bp{{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;background:#d1fae5;color:#065f46}}
.bf{{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;background:#fee2e2;color:#991b1b}}
.bw{{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;background:#fef3c7;color:#92400e}}
.bm{{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;background:#f1f5f9;color:#475569}}
.bn{{display:inline-block;padding:1px 6px;border-radius:3px;font-size:10px;font-weight:600;background:#dbeafe;color:#1e40af;margin-left:4px}}
.ec{{background:var(--c);border:1px solid var(--bd);border-radius:8px;margin-bottom:12px;overflow:hidden}}
.eh{{padding:14px 18px;cursor:pointer;display:flex;justify-content:space-between;align-items:center;user-select:none}}
.eh:hover{{background:#f8fafc}}
.et{{font-weight:600;font-size:14px}}
.et code{{background:#f1f5f9;padding:2px 6px;border-radius:4px;font-size:13px;margin-right:6px}}
.er{{display:flex;gap:8px;align-items:center}}
.arrow{{font-size:10px;color:var(--t2);transition:transform .2s}}
.eb{{padding:18px;display:none;border-top:1px solid var(--bd)}}
.ti{{background:#eff6ff;border:1px solid #bfdbfe;border-radius:6px;padding:10px 14px;margin-bottom:14px;font-size:13px;color:#1e40af}}
.ft{{width:100%;border-collapse:collapse;margin-bottom:14px;font-size:12px}}
.ft th{{background:#f8fafc;padding:7px 10px;text-align:left;font-weight:600;border-bottom:2px solid var(--bd)}}
.ft td{{padding:6px 10px;border-bottom:1px solid var(--bd);vertical-align:top}}
.ft .hl{{color:#dc2626;font-weight:600}}
.ev{{display:inline-block;padding:1px 5px;border-radius:3px;font-size:11px;margin:1px 2px}}
.ev.hit{{background:#d1fae5;color:#065f46}}
.ev.miss{{background:#fee2e2;color:#991b1b}}
.ev.act{{background:#e0e7ff;color:#3730a3}}
.jc{{background:#1e293b;border-radius:6px;padding:14px;overflow-x:auto;margin-top:10px;position:relative}}
.jc pre{{margin:0;font-family:'SF Mono','Fira Code','Consolas',monospace;font-size:12px;line-height:1.5;color:#e2e8f0;white-space:pre}}
.jk{{color:#93c5fd}}.js{{color:#86efac}}.jn{{color:#fcd34d}}.jb{{color:#c4b5fd}}.jnl{{color:#94a3b8}}
.jhl{{color:#fca5a5!important;font-weight:700}}
.jhv{{color:#fca5a5!important;font-weight:700;background:rgba(239,68,68,.12);padding:0 2px;border-radius:2px}}
.combo-label{{font-size:12px;color:var(--t2);margin:10px 0 4px;font-weight:600}}
.cnt{{font-size:12px;color:var(--t2);margin-top:6px;font-style:italic}}
.footer{{text-align:center;color:var(--t2);font-size:12px;margin-top:32px;padding-top:16px;border-top:1px solid var(--bd)}}
.bi{{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;background:#e0e7ff;color:#3730a3}}
.btn-ign{{border:1px solid var(--bd);background:var(--c);color:var(--t2);font-size:11px;padding:2px 8px;border-radius:4px;cursor:pointer}}
.btn-ign:hover{{background:#f1f5f9}}
.tr-ignored td{{opacity:.5;text-decoration:line-through}}
.sfr td{{padding-left:28px!important;font-size:11px;background:#fafbfc}}
.sfr td:first-child{{color:var(--t2)}}
</style>
</head>
<body>
<div class="ct">
<h1>{esc(title)}</h1>
<div class="meta">
<span>验证时间: {now}</span>
<span>数据来源: {esc(source_desc)}</span>
<span>匹配记录: {total_records} 条</span>
</div>

<div class="sm">
<div class="sc"><div class="lb">总事件数</div><div class="vl">{total}</div></div>
<div class="sc sp"><div class="lb">通过</div><div class="vl" id="sm-pass">{passed}</div></div>
<div class="sc sf"><div class="lb">异常</div><div class="vl" id="sm-fail">{failed}</div></div>
<div class="sc sw"><div class="lb">未触发</div><div class="vl" id="sm-nt">{not_triggered}</div></div>
</div>

<h2>概览</h2>
<table class="ot">
<tr><th>#</th><th>事件名</th><th>中文名</th><th>类型</th><th>状态</th><th>上报次数</th><th>取值种数</th><th>问题字段</th></tr>''']

    for i, r in enumerate(results, 1):
        ct = "修改" if r["change_type"] == "modify" else "新增" if r["change_type"] == "new" else "-"
        badge = status_badge(r["overall"])
        cnt = r["count"]
        combos = len(r["combinations"])
        issues = [f["field_id"] for f in r["fields"] if f["status"] not in ("pass",)]
        for f in r["fields"]:
            for sf in f.get("sub_field_results", []):
                if sf["status"] != "pass":
                    issues.append(f'{f["field_id"]}.{sf["field_id"]}')
        issue_str = ", ".join(issues) if issues else "-"
        html.append(f'<tr data-evt="{esc(r["event_name"])}" data-overall="{r["overall"]}"><td>{i}</td><td><code>{esc(r["event_name"])}</code></td><td>{esc(r["cn_name"])}</td><td>{ct}</td><td class="ot-status">{badge}</td><td>{cnt}</td><td>{combos}</td><td class="ot-issues">{esc(issue_str)}</td></tr>')

    html.append('</table>')

    # Event detail cards
    html.append('<h2>详细报告</h2>')
    for idx, r in enumerate(results):
        evt = r["event_name"]
        cn = r["cn_name"]
        badge = status_badge(r["overall"])
        cnt_label = f'{r["count"]}条' if r["triggered"] else "0条"
        hl_keys = r["doc_field_ids"]

        html.append(f'''<div class="ec" data-evt="{esc(evt)}">
<div class="eh" onclick="toggle('eb{idx}',this)">
<div class="et"><code>{esc(evt)}</code> {esc(cn)}</div>
<div class="er"><span class="evt-badge">{badge}</span> <span style="font-size:12px;color:var(--t2)">{cnt_label}</span> <span class="arrow">▼</span></div>
</div>
<div class="eb" id="eb{idx}" style="display:none">''')

        # Trigger info
        html.append(f'<div class="ti">触发行为: {esc(r["trigger_desc"])}</div>')

        if not r["triggered"]:
            html.append('<p style="color:var(--t2);font-size:13px">该事件在监听期间未被触发，请确认操作是否覆盖到相关场景。</p>')
            if r["fields"]:
                html.append('<p style="font-size:13px;margin-top:8px">文档要求字段：')
                for f in r["fields"]:
                    new_tag = '<span class="bn">新增</span>' if f.get("is_new") else ""
                    html.append(f'<code>{esc(f["field_id"])}</code>{new_tag} ')
                html.append('</p>')
        else:
            # Field coverage table
            if r["fields"]:
                html.append(f'''<table class="ft" data-evt="{esc(evt)}">
<tr><th>字段ID</th><th>字段名</th><th>类型</th><th>预期值</th><th>实际值</th><th>状态</th><th></th></tr>''')
                for f in r["fields"]:
                    fid_cls = ' class="hl"' if True else ""  # all documented fields highlighted
                    new_tag = ' <span class="bn">新增</span>' if f.get("is_new") else ""

                    # Expected values display
                    if f["expected_values"]:
                        ev_tags = []
                        for ev in f["expected_values"]:
                            hit = f["coverage"].get(ev, False)
                            cls = "ev hit" if hit else "ev miss"
                            ev_tags.append(f'<span class="{cls}">{esc(str(ev))}</span>')
                        ev_html = " ".join(ev_tags)
                    else:
                        ev_html = '<span style="color:var(--t2)">-</span>'

                    # Actual values display
                    if f["actual_values"]:
                        av_tags = []
                        for av in f["actual_values"][:8]:
                            s = str(av)
                            if len(s) > 30:
                                s = s[:30] + "..."
                            av_tags.append(f'<span class="ev act">{esc(s)}</span>')
                        if len(f["actual_values"]) > 8:
                            av_tags.append(f'<span style="font-size:11px;color:var(--t2)">等{len(f["actual_values"])}种</span>')
                        av_html = " ".join(av_tags)
                    else:
                        av_html = '<span style="color:var(--t2)">-</span>'

                    badge_html = status_badge(f["status"])
                    ign_btn = f'<button class="btn-ign" onclick="ignoreField(this,\'{esc(evt)}\')">忽略</button>' if f["status"] != "pass" else ""
                    tr_cls = f' data-status="{f["status"]}"'
                    html.append(f'<tr{tr_cls}><td{fid_cls}>{esc(f["field_id"])}{new_tag}</td><td>{esc(f["name"])}</td><td>{esc(f["type"])}</td><td>{ev_html}</td><td>{av_html}</td><td class="f-badge">{badge_html}</td><td>{ign_btn}</td></tr>')

                    # Render sub-field rows for JSON-type fields
                    for sf in f.get("sub_field_results", []):
                        sf_new_tag = ' <span class="bn">新增</span>' if sf.get("is_new") else ""
                        if sf["expected_values"]:
                            sf_ev_tags = []
                            for ev in sf["expected_values"]:
                                hit = sf["coverage"].get(ev, False)
                                cls = "ev hit" if hit else "ev miss"
                                sf_ev_tags.append(f'<span class="{cls}">{esc(str(ev))}</span>')
                            sf_ev_html = " ".join(sf_ev_tags)
                        else:
                            sf_ev_html = '<span style="color:var(--t2)">-</span>'
                        if sf["actual_values"]:
                            sf_av_tags = []
                            for av in sf["actual_values"][:8]:
                                s = str(av)
                                if len(s) > 30:
                                    s = s[:30] + "..."
                                sf_av_tags.append(f'<span class="ev act">{esc(s)}</span>')
                            if len(sf["actual_values"]) > 8:
                                sf_av_tags.append(f'<span style="font-size:11px;color:var(--t2)">等{len(sf["actual_values"])}种</span>')
                            sf_av_html = " ".join(sf_av_tags)
                        else:
                            sf_av_html = '<span style="color:var(--t2)">-</span>'
                        sf_badge_html = status_badge(sf["status"])
                        sf_ign_btn = f'<button class="btn-ign" onclick="ignoreField(this,\'{esc(evt)}\')">忽略</button>' if sf["status"] != "pass" else ""
                        sf_tr_cls = f' data-status="{sf["status"]}"'
                        html.append(f'<tr class="sfr"{sf_tr_cls}><td>└ {esc(sf["field_id"])}{sf_new_tag}</td><td>{esc(sf["name"])}</td><td>{esc(sf.get("type", ""))}</td><td>{sf_ev_html}</td><td>{sf_av_html}</td><td class="f-badge">{sf_badge_html}</td><td>{sf_ign_btn}</td></tr>')
                html.append('</table>')

            # JSON records
            combos = r["combinations"]
            max_show = 10
            if combos:
                if len(combos) == 1:
                    html.append('<div class="combo-label">实际上报内容：</div>')
                    html.append(f'<div class="jc"><pre>{render_json_html(combos[0], hl_keys)}</pre></div>')
                else:
                    for ci, combo in enumerate(combos[:max_show], 1):
                        html.append(f'<div class="combo-label">取值 {ci}/{len(combos)}：</div>')
                        html.append(f'<div class="jc"><pre>{render_json_html(combo, hl_keys)}</pre></div>')
                    if len(combos) > max_show:
                        html.append(f'<div class="cnt">还有 {len(combos) - max_show} 种取值未展示</div>')

        html.append('</div></div>')

    html.append(f'''
<div class="footer">dm-check 埋点验证报告 · 生成于 {now}</div>
</div>
<script>
function toggle(id, hdr) {{
  var el = document.getElementById(id);
  var arrow = hdr.querySelector('.arrow');
  if (el.style.display === 'none') {{
    el.style.display = 'block';
    arrow.textContent = '▼';
  }} else {{
    el.style.display = 'none';
    arrow.textContent = '▶';
  }}
}}

function ignoreField(btn, evtName) {{
  var tr = btn.closest('tr');
  tr.classList.add('tr-ignored');
  tr.setAttribute('data-status', 'ignored');
  tr.querySelector('.f-badge').innerHTML = '<span class="bi">已忽略</span>';
  btn.remove();
  refreshEvent(evtName);
  refreshSummary();
}}

function refreshEvent(evtName) {{
  // Check if all non-pass fields in this event are now ignored
  var card = document.querySelector('.ec[data-evt="' + evtName + '"]');
  if (!card) return;
  var rows = card.querySelectorAll('.ft tr[data-status]');
  var hasIssue = false;
  rows.forEach(function(r) {{
    var st = r.getAttribute('data-status');
    if (st && st !== 'pass' && st !== 'ignored') hasIssue = true;
  }});
  var newOverall = hasIssue ? 'has_issues' : 'pass';
  var badgeEl = card.querySelector('.evt-badge');
  if (badgeEl) {{
    if (newOverall === 'pass') {{
      badgeEl.innerHTML = '<span class="bp">通过</span>';
    }}
  }}
  // Update overview table row
  var otRow = document.querySelector('.ot tr[data-evt="' + evtName + '"]');
  if (otRow) {{
    otRow.setAttribute('data-overall', newOverall);
    var statusTd = otRow.querySelector('.ot-status');
    if (statusTd && newOverall === 'pass') {{
      statusTd.innerHTML = '<span class="bp">通过</span>';
    }}
    // Update issue fields column
    var issueParts = [];
    rows.forEach(function(r) {{
      var st = r.getAttribute('data-status');
      if (st && st !== 'pass' && st !== 'ignored') {{
        issueParts.push(r.querySelector('td').textContent.trim());
      }}
    }});
    var issueTd = otRow.querySelector('.ot-issues');
    if (issueTd) issueTd.textContent = issueParts.length ? issueParts.join(', ') : '-';
  }}
}}

function refreshSummary() {{
  var rows = document.querySelectorAll('.ot tr[data-overall]');
  var p = 0, f = 0, nt = 0;
  rows.forEach(function(r) {{
    var o = r.getAttribute('data-overall');
    if (o === 'pass') p++;
    else if (o === 'has_issues') f++;
    else if (o === 'not_triggered') nt++;
  }});
  var ep = document.getElementById('sm-pass');
  var ef = document.getElementById('sm-fail');
  var en = document.getElementById('sm-nt');
  if (ep) ep.textContent = p;
  if (ef) ef.textContent = f;
  if (en) en.textContent = nt;
}}

document.addEventListener('DOMContentLoaded', function() {{
  document.querySelectorAll('.ec').forEach(function(card) {{
    var badge = card.querySelector('.evt-badge');
    if (badge && badge.querySelector('.bf')) {{
      var body = card.querySelector('.eb');
      if (body) body.style.display = 'block';
    }}
  }});
}});
</script>
</body>
</html>''')

    report = '\n'.join(html)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(report)
    return report


# ==================== Main ====================

def main():
    parser = argparse.ArgumentParser(description="埋点验证工具 v2")
    parser.add_argument("--spec", help="Spec JSON 文件路径")
    parser.add_argument("--excel", help="埋点文档 Excel 文件路径")
    parser.add_argument("--source", choices=["elk", "logcat", "logcat-file"], default="logcat",
                        help="数据来源")
    parser.add_argument("--uid", help="用户 UID (ELK 模式)")
    parser.add_argument("--time-range", default="today", help="时间范围 (ELK 模式)")
    parser.add_argument("--logcat-file", help="logcat 文件路径")
    parser.add_argument("--sheets", help="Sheet 名称，逗号分隔")
    parser.add_argument("--events", help="事件名，逗号分隔")
    parser.add_argument("--output", default="/tmp/dm-check-report.html", help="HTML 报告输出路径")
    parser.add_argument("--parse-only", action="store_true", help="仅解析 Excel")
    parser.add_argument("--format", choices=["text", "json"], default="text", help="--parse-only 输出格式")
    args = parser.parse_args()

    if not args.spec and not args.excel:
        print("ERROR: --spec or --excel is required", file=sys.stderr)
        sys.exit(1)

    target_sheets = args.sheets.split(",") if args.sheets else None
    target_events = set(args.events.split(",")) if args.events else None

    # === Parse-only mode (Excel) ===
    if args.parse_only:
        if not args.excel:
            print("ERROR: --excel is required for --parse-only", file=sys.stderr)
            sys.exit(1)
        if not os.path.exists(args.excel):
            print(f"ERROR: Excel file not found: {args.excel}", file=sys.stderr)
            sys.exit(1)

        print("Parsing Excel...", file=sys.stderr)
        excel_events = parse_excel(args.excel, target_sheets)
        if target_events:
            excel_events = {k: v for k, v in excel_events.items() if k in target_events}
        print(f"  Found {len(excel_events)} events", file=sys.stderr)

        parsed = excel_to_parse_json(excel_events)

        if args.format == "json":
            print(json.dumps(parsed, ensure_ascii=False, indent=2))
        else:
            for evt_name, evt_info in parsed["events"].items():
                cn = evt_info.get("cn_name", "")
                ct = evt_info.get("change_type", "")
                tr = evt_info.get("trigger_rule", "")
                print(f"\n### {evt_name}（{cn}）[{ct}]")
                if tr:
                    print(f"  触发规则: {tr}")
                for fid, fdef in evt_info.get("fields", {}).items():
                    new_tag = " [本期新增]" if fdef.get("is_new") else ""
                    ev = fdef.get("extracted_values", [])
                    ev_str = f" → 预期值: {ev}" if ev else ""
                    print(f"  - {fid} ({fdef.get('name', '')}) [{fdef.get('type', '')}]{new_tag}{ev_str}")
                    if fdef.get("remark"):
                        remark_short = fdef["remark"].replace("\n", " | ")
                        if len(remark_short) > 100:
                            remark_short = remark_short[:100] + "..."
                        print(f"    备注: {remark_short}")
        return

    # === Load spec ===
    if args.spec:
        if not os.path.exists(args.spec):
            print(f"ERROR: Spec file not found: {args.spec}", file=sys.stderr)
            sys.exit(1)
        spec = load_spec(args.spec)
        print(f"Loaded spec with {len(spec.get('events', {}))} events", file=sys.stderr)
    else:
        # Excel → spec conversion
        if not os.path.exists(args.excel):
            print(f"ERROR: Excel file not found: {args.excel}", file=sys.stderr)
            sys.exit(1)
        print("Parsing Excel...", file=sys.stderr)
        excel_events = parse_excel(args.excel, target_sheets)
        if target_events:
            excel_events = {k: v for k, v in excel_events.items() if k in target_events}
        spec = excel_to_parse_json(excel_events)
        print(f"  Converted {len(spec.get('events', {}))} events to spec", file=sys.stderr)

    if target_events:
        spec["events"] = {k: v for k, v in spec.get("events", {}).items() if k in target_events}

    # === Fetch records ===
    source_desc = ""
    records = []

    if args.source == "elk":
        if not args.uid:
            print("ERROR: --uid is required for ELK mode", file=sys.stderr)
            sys.exit(1)
        start_ms, end_ms = parse_time_range(args.time_range)
        start_dt = datetime.fromtimestamp(start_ms / 1000, TZ_EAST8).strftime("%Y-%m-%d %H:%M:%S")
        end_dt = datetime.fromtimestamp(end_ms / 1000, TZ_EAST8).strftime("%Y-%m-%d %H:%M:%S")
        source_desc = f"ELK | UID: {args.uid} | {start_dt} ~ {end_dt}"
        print(f"Querying ELK ({start_dt} ~ {end_dt})...", file=sys.stderr)
        event_names = list(spec.get("events", {}).keys())
        hits = query_elk(args.uid, event_names, start_ms, end_ms)
        records = parse_elk_records(hits)
    elif args.source == "logcat":
        source_desc = "adb logcat (设备缓冲区)"
        print("Capturing logcat from device...", file=sys.stderr)
        records = capture_logcat()
    elif args.source == "logcat-file":
        if not args.logcat_file or not os.path.exists(args.logcat_file):
            print(f"ERROR: logcat file not found: {args.logcat_file}", file=sys.stderr)
            sys.exit(1)
        source_desc = f"logcat 文件: {os.path.basename(args.logcat_file)}"
        print(f"Reading logcat file: {args.logcat_file}...", file=sys.stderr)
        records = read_logcat(args.logcat_file)

    print(f"  Parsed {len(records)} valid records", file=sys.stderr)

    # === Verify ===
    print("Verifying...", file=sys.stderr)
    results = verify(spec, records)

    # === Generate HTML ===
    generate_html(results, spec, source_desc, len(records), args.output)
    print(f"Report saved to: {args.output}", file=sys.stderr)

    # Print summary to stdout
    total = len(results)
    passed = sum(1 for r in results if r["overall"] == "pass")
    failed = sum(1 for r in results if r["overall"] == "has_issues")
    not_triggered = sum(1 for r in results if r["overall"] == "not_triggered")
    print(f"\n验证完成: {total}个事件 | 通过 {passed} | 异常 {failed} | 未触发 {not_triggered}")
    print(f"HTML 报告: {args.output}")


if __name__ == "__main__":
    main()
